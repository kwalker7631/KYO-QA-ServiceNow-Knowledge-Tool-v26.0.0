# processing_engine.py - Fixed version with robust error handling
import shutil
import time
import json
import openpyxl
import re
from queue import Queue
from pathlib import Path
from datetime import datetime
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import *
from custom_exceptions import FileLockError
from data_harvesters import harvest_all_data
from file_utils import is_file_locked
from ocr_utils import extract_text_from_pdf, _is_ocr_needed

def clear_review_folder():
    """Clear the review folder of any previous text files"""
    try:
        if PDF_TXT_DIR.exists():
            review_folder = PDF_TXT_DIR / "needs_review"
            review_folder.mkdir(exist_ok=True)
            
            for f in review_folder.glob("*.txt"):
                try:
                    f.unlink()
                except OSError as e:
                    print(f"Warning: Could not delete review file {f}: {e}")
        else:
            PDF_TXT_DIR.mkdir(parents=True, exist_ok=True)
            (PDF_TXT_DIR / "needs_review").mkdir(exist_ok=True)
    except Exception as e:
        print(f"Error clearing review folder: {e}")

def get_cache_path(pdf_path):
    """Generate cache file path for a PDF"""
    try:
        pdf_path = Path(pdf_path)
        file_size = pdf_path.stat().st_size if pdf_path.exists() else 0
        cache_name = f"{pdf_path.stem}_{file_size}.json"
        return CACHE_DIR / cache_name
    except Exception as e:
        print(f"Error generating cache path for {pdf_path}: {e}")
        return CACHE_DIR / f"{Path(pdf_path).stem}_unknown.json"

def load_cached_result(cache_path):
    """Load cached processing result if available"""
    try:
        if not cache_path.exists():
            return None
        
        with open(cache_path, 'r', encoding='utf-8') as f:
            cached_data = json.load(f)
        
        # Validate cached data structure
        required_keys = ["filename", "models", "status"]
        if not all(key in cached_data for key in required_keys):
            print(f"Invalid cache format in {cache_path}")
            return None
        
        return cached_data
        
    except (json.JSONDecodeError, FileNotFoundError, KeyError) as e:
        print(f"Could not load cache from {cache_path}: {e}")
        return None
    except Exception as e:
        print(f"Unexpected error loading cache: {e}")
        return None

def save_cached_result(cache_path, result_data):
    """Save processing result to cache"""
    try:
        # Ensure cache directory exists
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(result_data, f, indent=2, ensure_ascii=False)
            
    except Exception as e:
        print(f"Could not save cache to {cache_path}: {e}")

def process_single_pdf(pdf_path, progress_queue, ignore_cache=False):
    """
    Process a single PDF file and extract model information.
    Returns a dictionary with the processing results.
    """
    try:
        # Ensure pdf_path is a Path object for consistency
        pdf_path = Path(pdf_path)
        filename = pdf_path.name
        
        # Generate cache path
        cache_path = get_cache_path(pdf_path)

        # Log which file we're processing
        progress_queue.put({
            "type": "log", 
            "tag": "info", 
            "msg": f"Processing: {filename}"
        })
        
        # Try to load from cache first (unless ignoring cache)
        if not ignore_cache:
            cached_result = load_cached_result(cache_path)
            if cached_result:
                progress_queue.put({
                    "type": "log", 
                    "tag": "info", 
                    "msg": f"Loaded from cache: {filename}"
                })
                
                # Handle cached review items
                if cached_result.get("status") == "Needs Review":
                    review_info = cached_result.get("review_info")
                    if review_info:
                        progress_queue.put({
                            "type": "review_item", 
                            "data": review_info
                        })
                
                # Update counters
                progress_queue.put({
                    "type": "file_complete", 
                    "status": cached_result.get("status", "Unknown")
                })
                
                if cached_result.get("ocr_used"):
                    progress_queue.put({
                        "type": "increment_counter", 
                        "counter": "ocr"
                    })
                
                return cached_result

        # Update status
        progress_queue.put({
            "type": "status", 
            "msg": filename, 
            "led": "Queued"
        })
        
        # Check if file exists and is accessible
        if not pdf_path.exists():
            error_result = {
                "filename": filename,
                "models": "Error: File Not Found",
                "author": "",
                "status": "Fail",
                "ocr_used": False,
                "review_info": None,
                "error": "File not found"
            }
            save_cached_result(cache_path, error_result)
            return error_result
        
        # Check if OCR will be needed
        absolute_pdf_path = str(pdf_path.resolve())
        ocr_required = False
        
        try:
            ocr_required = _is_ocr_needed(absolute_pdf_path)
            if ocr_required:
                progress_queue.put({
                    "type": "status", 
                    "msg": filename, 
                    "led": "OCR"
                })
                progress_queue.put({
                    "type": "increment_counter", 
                    "counter": "ocr"
                })
        except Exception as e:
            print(f"Error checking OCR requirements for {filename}: {e}")
        
        # Extract text from PDF
        progress_queue.put({
            "type": "status", 
            "msg": f"Extracting text: {filename}", 
            "led": "Processing"
        })
        
        extracted_text = extract_text_from_pdf(absolute_pdf_path)
        
        # Check if text extraction was successful
        if not extracted_text or len(extracted_text.strip()) < 10:
            error_result = {
                "filename": filename,
                "models": "Error: Text Extraction Failed",
                "author": "",
                "status": "Fail",
                "ocr_used": ocr_required,
                "review_info": None,
                "error": "No text could be extracted"
            }
            save_cached_result(cache_path, error_result)
            progress_queue.put({
                "type": "file_complete", 
                "status": "Fail"
            })
            return error_result
        
        # Process the extracted text
        progress_queue.put({
            "type": "status", 
            "msg": f"Analyzing content: {filename}", 
            "led": "AI"
        })
        
        extraction_data = harvest_all_data(extracted_text, filename)
        
        # Determine processing status
        if extraction_data["models"] == "Not Found" or not extraction_data["models"]:
            status = "Needs Review"
            
            # Create review file
            review_folder = PDF_TXT_DIR / "needs_review"
            review_folder.mkdir(parents=True, exist_ok=True)
            review_txt_path = review_folder / f"{pdf_path.stem}.txt"
            
            try:
                with open(review_txt_path, 'w', encoding='utf-8') as f:
                    f.write(f"--- Filename: {filename} ---\n")
                    f.write(f"--- Processing Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ---\n")
                    f.write(f"--- Extraction Stats: {extraction_data.get('extraction_stats', {})} ---\n\n")
                    f.write(extracted_text)
                
                review_info = {
                    "filename": filename,
                    "reason": "No models found",
                    "txt_path": str(review_txt_path),
                    "pdf_path": str(pdf_path),
                    "extraction_stats": extraction_data.get('extraction_stats', {})
                }
                
                progress_queue.put({
                    "type": "review_item", 
                    "data": review_info
                })
                
            except Exception as e:
                print(f"Error creating review file for {filename}: {e}")
                review_info = None
                
        else:
            status = "Pass"
            review_info = None
        
        # Create result dictionary
        result = {
            "filename": filename,
            "models": extraction_data["models"],
            "author": extraction_data.get("author", ""),
            "qa_numbers": extraction_data.get("qa_numbers", ""),
            "description": extraction_data.get("description", ""),
            "status": status,
            "ocr_used": ocr_required,
            "review_info": review_info,
            "extraction_stats": extraction_data.get("extraction_stats", {}),
            "processed_date": datetime.now().isoformat()
        }
        
        # Save to cache
        save_cached_result(cache_path, result)
        
        # Update progress
        progress_queue.put({
            "type": "file_complete", 
            "status": result["status"]
        })
        
        return result
        
    except Exception as e:
        # Handle unexpected errors
        error_result = {
            "filename": Path(pdf_path).name,
            "models": f"Error: Processing Failed ({str(e)[:50]})",
            "author": "",
            "status": "Fail",
            "ocr_used": False,
            "review_info": None,
            "error": str(e)
        }
        
        print(f"Error processing {pdf_path}: {e}")
        
        try:
            cache_path = get_cache_path(pdf_path)
            save_cached_result(cache_path, error_result)
        except:
            pass  # Don't let cache saving errors crash the process
        
        progress_queue.put({
            "type": "file_complete", 
            "status": "Fail"
        })
        
        return error_result

def update_excel_file(excel_path, results_dict, progress_queue):
    """Update Excel file with processing results"""
    try:
        progress_queue.put({
            "type": "status", 
            "msg": "Loading Excel file...", 
            "led": "Saving"
        })
        
        # Load workbook
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        
        # Get headers
        headers = [cell.value for cell in sheet[1]]
        
        # Add processing status column if it doesn't exist
        if STATUS_COLUMN_NAME not in headers:
            new_col = len(headers) + 1
            sheet.cell(row=1, column=new_col).value = STATUS_COLUMN_NAME
            headers.append(STATUS_COLUMN_NAME)
        
        # Create column mapping
        column_mapping = {}
        for col_name in [DESCRIPTION_COLUMN_NAME, META_COLUMN_NAME, AUTHOR_COLUMN_NAME, STATUS_COLUMN_NAME]:
            if col_name in headers:
                column_mapping[col_name] = headers.index(col_name) + 1
        
        progress_queue.put({
            "type": "status", 
            "msg": "Updating Excel rows...", 
            "led": "Saving"
        })
        
        # Update rows based on matching filenames
        rows_updated = 0
        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            desc_cell = row[column_mapping.get(DESCRIPTION_COLUMN_NAME, 1) - 1]
            desc_value = str(desc_cell.value) if desc_cell.value else ""
            
            # Find matching result
            matched_result = None
            for filename, result_data in results_dict.items():
                filename_stem = Path(filename).stem
                if filename_stem in desc_value:
                    matched_result = result_data
                    break
            
            if matched_result:
                # Update Meta column
                if META_COLUMN_NAME in column_mapping:
                    meta_cell = row[column_mapping[META_COLUMN_NAME] - 1]
                    if not meta_cell.value or str(meta_cell.value).strip() == "":
                        meta_cell.value = matched_result["models"]
                
                # Update Author column
                if AUTHOR_COLUMN_NAME in column_mapping and matched_result.get("author"):
                    author_cell = row[column_mapping[AUTHOR_COLUMN_NAME] - 1]
                    if not author_cell.value or str(author_cell.value).strip() == "":
                        author_cell.value = matched_result["author"]
                
                # Update Status column
                if STATUS_COLUMN_NAME in column_mapping:
                    status_cell = row[column_mapping[STATUS_COLUMN_NAME] - 1]
                    status_text = matched_result["status"]
                    if matched_result.get("ocr_used"):
                        status_text += " (OCR)"
                    status_cell.value = status_text
                
                rows_updated += 1
        
        progress_queue.put({
            "type": "status", 
            "msg": "Applying formatting...", 
            "led": "Saving"
        })
        
        # Apply conditional formatting
        apply_excel_formatting(sheet, column_mapping.get(STATUS_COLUMN_NAME))
        
        # Auto-adjust column widths
        adjust_column_widths(sheet)
        
        # Save workbook
        workbook.save(excel_path)
        
        progress_queue.put({
            "type": "log", 
            "tag": "success", 
            "msg": f"Excel updated successfully: {rows_updated} rows modified"
        })
        
        return True
        
    except Exception as e:
        progress_queue.put({
            "type": "log", 
            "tag": "error", 
            "msg": f"Error updating Excel file: {e}"
        })
        return False

def apply_excel_formatting(sheet, status_column_index):
    """Apply color-coded formatting based on processing status"""
    try:
        if not status_column_index:
            return
        
        # Define fill colors
        fills = {
            "Pass": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Light green
            "Fail": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # Light red
            "Needs Review": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # Light yellow
            "OCR": PatternFill(start_color="0A9BCD", end_color="0A9BCD", fill_type="solid")  # Light blue
        }
        
        # Apply formatting to data rows
        for row in sheet.iter_rows(min_row=2):
            status_cell = row[status_column_index - 1]
            status_value = str(status_cell.value) if status_cell.value else ""
            
            # Determine fill color
            fill_color = None
            for status_key, fill in fills.items():
                if status_key in status_value:
                    fill_color = fill
                    break
            
            # Apply fill to entire row if status found
            if fill_color:
                for cell in row:
                    cell.fill = fill_color
            
            # Special handling for OCR indicator
            if "(OCR)" in status_value:
                status_cell.fill = fills["OCR"]
                
    except Exception as e:
        print(f"Error applying Excel formatting: {e}")

def adjust_column_widths(sheet):
    """Auto-adjust column widths for better readability"""
    try:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Set width with reasonable limits
            adjusted_width = min(max(max_length + 2, 10), 70)
            sheet.column_dimensions[column_letter].width = adjusted_width
            
    except Exception as e:
        print(f"Error adjusting column widths: {e}")

def run_processing_job(job_info, progress_queue, cancel_event, pause_event):
    """
    Main processing job coordinator with robust error handling.
    """
    try:
        # Extract job parameters
        is_rerun = job_info.get("is_rerun", False)
        excel_path = Path(job_info["excel_path"])
        input_path = job_info["input_path"]
        
        progress_queue.put({
            "type": "log", 
            "tag": "info", 
            "msg": f"Processing job started ({'rerun' if is_rerun else 'new'})"
        })
        
        # Handle Excel file cloning or reuse
        if is_rerun:
            # For rerun, use the existing cloned file
            clear_review_folder()
            cloned_path = excel_path
        else:
            # For new job, create a clone
            timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
            cloned_path = OUTPUT_DIR / f"cloned_{excel_path.stem}_{timestamp}{excel_path.suffix}"
            
            # Check if input file is locked
            if is_file_locked(excel_path):
                raise FileLockError(f"Input Excel file is locked: {excel_path}")
            
            # Create output directory if needed
            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            
            # Copy the Excel file
            shutil.copy(excel_path, cloned_path)
            progress_queue.put({
                "type": "log", 
                "tag": "info", 
                "msg": f"Created Excel copy: {cloned_path.name}"
            })
        
        # Determine files to process
        if isinstance(input_path, list):
            # List of specific files
            files_to_process = [Path(f) for f in input_path]
        else:
            # Directory of PDF files
            input_dir = Path(input_path)
            if not input_dir.exists():
                raise FileNotFoundError(f"Input directory not found: {input_dir}")
            
            files_to_process = list(input_dir.glob('*.pdf'))
        
        # Filter out non-existent files
        valid_files = [f for f in files_to_process if f.exists()]
        
        if not valid_files:
            progress_queue.put({
                "type": "log", 
                "tag": "warning", 
                "msg": "No valid PDF files found to process"
            })
            progress_queue.put({"type": "finish", "status": "No Files"})
            return
        
        progress_queue.put({
            "type": "log", 
            "tag": "info", 
            "msg": f"Found {len(valid_files)} PDF files to process"
        })
        
        # Process each file
        results = {}
        processed_count = 0
        
        for i, pdf_path in enumerate(valid_files):
            # Check for cancellation
            if cancel_event.is_set():
                progress_queue.put({
                    "type": "log", 
                    "tag": "warning", 
                    "msg": "Processing cancelled by user"
                })
                break
            
            # Check for pause
            while pause_event and pause_event.is_set():
                progress_queue.put({
                    "type": "status", 
                    "msg": "Paused", 
                    "led": "Paused"
                })
                time.sleep(0.5)
            
            # Update progress
            progress_queue.put({
                "type": "progress", 
                "current": i + 1, 
                "total": len(valid_files)
            })
            
            # Process the PDF
            try:
                result = process_single_pdf(pdf_path, progress_queue, ignore_cache=is_rerun)
                if result:
                    results[result["filename"]] = result
                    processed_count += 1
            except Exception as e:
                progress_queue.put({
                    "type": "log", 
                    "tag": "error", 
                    "msg": f"Failed to process {pdf_path.name}: {e}"
                })
        
        # Check if processing was cancelled
        if cancel_event.is_set():
            progress_queue.put({"type": "finish", "status": "Cancelled"})
            return
        
        # Update Excel file with results
        if results:
            success = update_excel_file(cloned_path, results, progress_queue)
            if success:
                progress_queue.put({
                    "type": "result_path", 
                    "path": str(cloned_path)
                })
            else:
                progress_queue.put({
                    "type": "log", 
                    "tag": "error", 
                    "msg": "Failed to update Excel file"
                })
        
        # Job completion
        completion_status = "Complete" if processed_count > 0 else "No Results"
        progress_queue.put({
            "type": "log", 
            "tag": "success", 
            "msg": f"Processing completed: {processed_count} files processed"
        })
        progress_queue.put({"type": "finish", "status": completion_status})
        
    except Exception as e:
        # Handle unexpected errors
        error_msg = f"Critical processing error: {e}"
        progress_queue.put({
            "type": "log", 
            "tag": "error", 
            "msg": error_msg
        })
        progress_queue.put({"type": "finish", "status": f"Error: {str(e)[:50]}"})
        
        # Print full traceback for debugging
        import traceback
        traceback.print_exc()

# Test functionality if run directly
if __name__ == "__main__":
    print("Processing Engine - Standalone Test")
    print("=" * 40)
    
    # This would be used for testing individual components
    test_pdf = Path("test.pdf")
    if test_pdf.exists():
        print(f"Testing with {test_pdf}")
        import queue
        test_queue = queue.Queue()
        result = process_single_pdf(test_pdf, test_queue)
        print(f"Result: {result}")
    else:
        print("No test.pdf found for testing")
