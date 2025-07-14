# excel_generator.py
# Version: 2.3.0
# Last modified: 2025-07-13

import openpyxl
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path
from typing import List, Dict, Any, Tuple

logger = logging.getLogger(__name__)

class ExcelGenerator:
    """Handles creation of the Excel report from a template."""

    def __init__(self, template_path: str):
        self.template_path = template_path

    def _get_column_map(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, int]:
        """Creates a mapping from header names to column indices."""
        return {str(cell.value).strip(): cell.column for cell in sheet[1] if cell.value}

    def _autosize_columns(self, sheet: openpyxl.worksheet.worksheet.Worksheet):
        """Automatically adjusts column widths to fit the content."""
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    def create_report(self, results: List[Dict[str, Any]], output_filename: str) -> Tuple[Path, List[str]]:
        """
        Populates the Excel report, handles nested meta data, autosizes columns, and saves the file.
        """
        try:
            workbook = openpyxl.load_workbook(self.template_path)
            sheet = workbook.active
            column_map = self._get_column_map(sheet)
            
            if not column_map:
                raise ValueError("Could not map headers in the Excel template.")

            skipped_files = []
            row_idx = 2

            for result in results:
                if result.get("status") == "Fail":
                    skipped_files.append(result.get("filename", "Unknown File"))
                    continue

                for key, value in result.items():
                    if key == 'meta' and isinstance(value, dict):
                        # Correctly format the nested meta dictionary into a string
                        meta_parts = []
                        for meta_key, meta_value in value.items():
                            if isinstance(meta_value, list):
                                meta_parts.append(f"{meta_key}: {', '.join(meta_value)}")
                            else:
                                meta_parts.append(f"{meta_key}: {meta_value}")
                        cell_value = "; ".join(meta_parts)
                    elif isinstance(value, list):
                        cell_value = ", ".join(value)
                    else:
                        cell_value = value
                    
                    if key in column_map:
                        sheet.cell(row=row_idx, column=column_map[key], value=cell_value)
                
                row_idx += 1

            # Autosize columns before saving
            self._autosize_columns(sheet)

            output_path = Path(output_filename)
            workbook.save(output_path)
            logger.info(f"Excel report successfully generated: {output_path}")
            return output_path, skipped_files

        except FileNotFoundError:
            logger.error(f"Template file not found at: {self.template_path}")
            raise
        except Exception as e:
            logger.error(f"An error occurred during Excel report generation: {e}", exc_info=True)
            raise
