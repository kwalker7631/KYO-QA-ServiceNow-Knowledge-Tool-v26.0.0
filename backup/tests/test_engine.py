import queue
from pathlib import Path
import sys
import types
from types import SimpleNamespace

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

dummy = types.SimpleNamespace()
for name in [
    "fitz",
    "pytesseract",
    "cv2",
    "anthropic",
    "sentry_sdk",
    "sentry_sdk.integrations.logging",
    "extract.common",
    "custom_recycles",
]:
    sys.modules.setdefault(name, dummy)
sys.modules.setdefault("PIL", types.SimpleNamespace(Image=types.SimpleNamespace()))

if "openpyxl" not in sys.modules:
    class _Cell:
        def __init__(self, value=None):
            self.value = value

    class _Sheet:
        def __init__(self):
            self.rows = []

        def append(self, row):
            self.rows.append(list(row))

        def __getitem__(self, idx):
            return [_Cell(v) for v in self.rows[idx - 1]]

        def cell(self, row, column):
            while len(self.rows) < row:
                self.rows.append([])
            row_list = self.rows[row - 1]
            while len(row_list) < column:
                row_list.append(None)

            class Proxy:
                @property
                def value(self):
                    return row_list[column - 1]

                @value.setter
                def value(self, v):
                    row_list[column - 1] = v

            return Proxy()

        def iter_cols(self, min_col, max_col, min_row):
            max_row = len(self.rows)
            for row in range(min_row, max_row + 1):
                val = None
                if row - 1 < len(self.rows) and min_col - 1 < len(self.rows[row - 1]):
                    val = self.rows[row - 1][min_col - 1]
                yield _Cell(val)

    class _Workbook:
        def __init__(self):
            self.active = _Sheet()

        def save(self, path):
            Path(path).touch()
            SAVED[str(path)] = self

    SAVED = {}

    def _load_workbook(path):
        return SAVED.get(str(path), _Workbook())

    openpyxl = types.SimpleNamespace(Workbook=_Workbook, load_workbook=_load_workbook)
    sys.modules.setdefault("openpyxl", openpyxl)
else:
    import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

dummy = SimpleNamespace()
for mod in [
    "fitz",
    "pytesseract",
    "cv2",
    "anthropic",
    "sentry_sdk",
    "sentry_sdk.integrations.logging",
    "extract.common",
    "custom_recycles",
]:
    sys.modules.setdefault(mod, dummy)
sys.modules.setdefault("PIL", SimpleNamespace(Image=SimpleNamespace()))

from processing_engine import export_to_excel


def test_export_to_excel(tmp_path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Number", "meta", "author"])
    sheet.append(["123", "", ""])
    template_path = tmp_path / "template.xlsx"
    wb.save(template_path)

    results = [
        {
            "filename": "123.pdf",
            "models": "Model1",
            "qa_numbers": "QA-001",
            "author": "John Doe",
            "qa_numbers": ["QA-1"],
            "status": "Pass",
        }
    ]

    output_path = export_to_excel(results, template_path, queue.Queue())

    assert output_path is not None
    assert Path(output_path).exists()

    out_wb = openpyxl.load_workbook(output_path)
    out_sheet = out_wb.active
    assert out_sheet.cell(row=2, column=2).value == "Model1, QA-1"
    assert out_sheet.cell(row=2, column=3).value == "John Doe"


def test_export_to_excel_with_qa_column(tmp_path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Number", "meta", "author", "QA Numbers"])
    sheet.append(["123", "", "", ""])
    template_path = tmp_path / "template.xlsx"
    wb.save(template_path)

    results = [
        {
            "filename": "123.pdf",
            "models": "Model1",
            "author": "John Doe",
            "qa_numbers": ["QA-1", "QA-2"],
            "status": "Pass",
        }
    ]

    output_path, skipped = export_to_excel(results, template_path, queue.Queue())

    assert output_path is not None
    assert Path(output_path).exists()
    assert skipped == []

    out_wb = openpyxl.load_workbook(output_path)
    out_sheet = out_wb.active
    assert out_sheet.cell(row=2, column=2).value == "Model1; QA-001"
    assert out_sheet.cell(row=2, column=3).value == "John Doe"

def test_export_to_excel_multiple_rows(tmp_path):
    """Ensure multiple rows are mapped correctly."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Number", "meta", "author"])
    sheet.append(["123", "", ""])
    template_path = tmp_path / "template.xlsx"
    wb.save(template_path)

    results = [
        {
            "filename": "123.pdf",
            "models": "Model1",
            "author": "John Doe",
            "status": "Pass",
        },
        {
            "filename": "456.pdf",
            "models": "Model2",
            "author": "Jane Roe",
            "status": "Pass",
        },
    ]

    output_path = export_to_excel(results, template_path, queue.Queue())
    out_wb = openpyxl.load_workbook(output_path)
    out_sheet = out_wb.active
    assert out_sheet.cell(row=2, column=2).value == "Model1"
    assert out_sheet.cell(row=3, column=2).value == "Model2"
 
